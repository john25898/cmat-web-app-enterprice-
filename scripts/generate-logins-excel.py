"""Generate an Excel file with logins for Facility In-Charges, County Reps and Program HR."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r"c:\Users\ADMIN\Downloads\cm-a-t-enterprise-web-app"
OUT = BASE + r"\CHAK_Managerial_Logins.xlsx"

with open(BASE + r"\data\users.json", encoding="utf-8") as f:
    users = json.load(f)

inch = sorted(
    [u for u in users.values() if u["role"] == "facility_incharge"],
    key=lambda u: (u.get("county", ""), u.get("facility", "")),
)
reps = sorted(
    [u for u in users.values() if u["role"] == "county_rep"],
    key=lambda u: u.get("county", ""),
)
hrs = [u for u in users.values() if u["role"] == "program_hr"]

# ---------------- styling helpers ----------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")

wb = Workbook()


def style_header(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_rows(ws, start_row, rows):
    r = start_row
    for rec in rows:
        for c, val in enumerate(rec, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = LEFT if c in (2, 3, 5) else CENTER
        r += 1
    return r


# ---------------- Sheet 1: Facility In-Charges ----------------
ws1 = wb.active
ws1.title = "Facility In-Charges"
set_widths(ws1, [6, 22, 46, 20, 42, 16])
ws1.cell(row=1, column=1, value="CHAK - FACILITY IN-CHARGE LOGINS").font = Font(
    bold=True, size=14, color="1F4E78"
)
ws1.merge_cells("A1:F1")
ws1.cell(row=2, column=1, value="Login at the web app with the County Rep / Facility In-Charge role").font = Font(
    italic=True, size=10, color="808080"
)
ws1.merge_cells("A2:F2")

hdr = 4
for c, h in enumerate(["No.", "Name", "Username (Email)", "Password", "Facility", "County"], start=1):
    ws1.cell(row=hdr, column=c, value=h)
style_header(ws1, hdr, range(1, 7))

# demo in-charge first (Main Facility)
demo = [u for u in inch if u.get("facility") == "Main Facility"]
real = [u for u in inch if u.get("facility") != "Main Facility"]

r = hdr + 1
if demo:
    ws1.cell(row=r, column=1, value="DEMO").font = Font(bold=True, color="C00000")
    write_rows(ws1, r, [[1, u["name"], u["email"], u["password"], u.get("facility", ""), u.get("county", "")] for u in demo])
    ws1.cell(row=r, column=1, value="DEMO").font = Font(bold=True, color="C00000")
    r += len(demo)

for i, u in enumerate(real, start=1):
    row = [i, u["name"], u["email"], u["password"], u.get("facility", ""), u.get("county", "")]
    for c, val in enumerate(row, start=1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = LEFT if c in (2, 3, 5) else CENTER
    r += 1

last = r - 1
for rr in range(hdr + 1, last + 1):
    ws1.cell(row=rr, column=1).number_format = "0"
ws1.freeze_panes = "A5"

# ---------------- Sheet 2: County Reps ----------------
ws2 = wb.create_sheet("County Reps")
set_widths(ws2, [6, 22, 46, 20, 16])
ws2.cell(row=1, column=1, value="CHAK - COUNTY REPRESENTATIVE LOGINS").font = Font(
    bold=True, size=14, color="1F4E78"
)
ws2.merge_cells("A1:E1")
ws2.cell(row=2, column=1, value="Each rep oversees one county's facility approvals").font = Font(
    italic=True, size=10, color="808080"
)
ws2.merge_cells("A2:E2")

for c, h in enumerate(["No.", "Name", "Username (Email)", "Password", "County"], start=1):
    ws2.cell(row=4, column=c, value=h)
style_header(ws2, 4, range(1, 6))

reps_real = [u for u in reps if not u["email"].startswith("countyrep@")]
reps_demo = [u for u in reps if u["email"].startswith("countyrep@")]

r = 5
if reps_demo:
    row = [1, reps_demo[0]["name"], reps_demo[0]["email"], reps_demo[0]["password"], reps_demo[0].get("county", "")]
    for c, val in enumerate(row, start=1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = LEFT if c in (2, 3) else CENTER
    ws2.cell(row=r, column=1, value="DEMO").font = Font(bold=True, color="C00000")
    r += 1

for i, u in enumerate(reps_real, start=1):
    row = [i, u["name"], u["email"], u["password"], u.get("county", "")]
    for c, val in enumerate(row, start=1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = LEFT if c in (2, 3) else CENTER
    r += 1

ws2.freeze_panes = "A5"

# ---------------- Sheet 3: Program HR ----------------
ws3 = wb.create_sheet("Program HR")
set_widths(ws3, [6, 22, 46, 20])
ws3.cell(row=1, column=1, value="CHAK - PROGRAM HR LOGIN").font = Font(
    bold=True, size=14, color="1F4E78"
)
ws3.merge_cells("A1:D1")

for c, h in enumerate(["No.", "Name", "Username (Email)", "Password"], start=1):
    ws3.cell(row=3, column=c, value=h)
style_header(ws3, 3, range(1, 5))

r = 4
for i, u in enumerate(hrs, start=1):
    row = [i, u["name"], u["email"], u["password"]]
    for c, val in enumerate(row, start=1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = LEFT if c in (2, 3) else CENTER
    r += 1

# ---------------- totals footer on sheet 1 ----------------
tot_row = last + 2
ws1.cell(row=tot_row, column=2, value=f"TOTAL FACILITY IN-CHARGES: {len(inch)}")
ws1.cell(row=tot_row, column=2).font = Font(bold=True, color="1F4E78")

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Facility In-Charges: {len(inch)} (incl. 1 demo)")
print(f"County Reps: {len(reps)} (incl. 1 demo)")
print(f"Program HR: {len(hrs)}")
