#!/usr/bin/env python3
"""Generate login credentials for all CHAK employees from Active_Employees_181.xlsx
and append Username/Password columns to the Excel file. Also emits data/users.json
for the web app.

Passwords are UNIQUE per account (random, secrets-based). Re-running the script
keeps existing passwords stable — it only generates new ones for accounts that
don't have one yet.
"""
import json
import os
import re
import secrets
import sys
import unicodedata
from collections import Counter

import openpyxl

XLSX = "Active_Employees_181.xlsx"
OUT_JSON = "data/users.json"

# Pass --force to regenerate ALL passwords (ignores existing values).
FORCE = "--force" in sys.argv

# Safe alphabet — excludes lookalike characters (0/O, 1/l)
_ALPHABET = "abcdefghjkmnpqrstuvwxyzABCDEFGHJKMNPQRSTUVWXYZ23456789"
_PREFIX = "Chak!"


def gen_password() -> str:
    return _PREFIX + "".join(secrets.choice(_ALPHABET) for _ in range(8))


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text or "")
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", ".", text)
    text = text.strip(".")
    return text


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    # Find column indexes (1-based)
    col_name = headers.index("Name") + 1
    col_phone = headers.index("Phone") + 1
    col_desig = headers.index("Designation") + 1
    col_county = headers.index("County") + 1
    col_facility = headers.index("Station/Facility") + 1
    col_id = headers.index("ID Number") + 1

    # Append Username / Password columns (idempotent — skip if already present)
    existing_headers = [str(c.value or "").strip() for c in ws[1]]
    if "Username" in existing_headers and "Password" in existing_headers:
        username_col = existing_headers.index("Username") + 1
        password_col = existing_headers.index("Password") + 1
        print("Credentials columns already present — updating values only.")
    else:
        username_col = ws.max_column + 1
        password_col = username_col + 1
        ws.cell(row=1, column=username_col, value="Username")
        ws.cell(row=1, column=password_col, value="Password")

    # Existing passwords already written to the Excel (so re-runs are stable)
    existing_excel_passwords: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2):
        email = row[username_col - 1].value
        pwd = row[password_col - 1].value
        if email and pwd:
            existing_excel_passwords[str(email).strip().lower()] = str(pwd).strip()

    # Existing passwords from the previous users.json (for supervisors/demo)
    existing_json_passwords: dict[str, str] = {}
    if os.path.exists(OUT_JSON):
        try:
            with open(OUT_JSON, "r", encoding="utf-8") as f:
                prev = json.load(f)
            for k, v in prev.items():
                if v.get("password"):
                    existing_json_passwords[k.strip().lower()] = v["password"]
        except Exception:
            pass

    users = {}
    used_usernames = Counter()
    used_passwords: set[str] = set()
    if not FORCE:
        used_passwords = set(existing_excel_passwords.values()) | set(
            existing_json_passwords.values()
        )

    for row in ws.iter_rows(min_row=2):
        name = row[col_name - 1].value
        if not name:
            continue
        name = str(name).strip()
        phone = str(row[col_phone - 1].value or "").strip()
        designation = str(row[col_desig - 1].value or "").strip()
        county = str(row[col_county - 1].value or "").strip()
        facility = str(row[col_facility - 1].value or "").strip()
        id_num = str(row[col_id - 1].value or "").strip()

        base = slugify(name)
        used_usernames[base] += 1
        username = base if used_usernames[base] == 1 else f"{base}.{used_usernames[base]}"
        email = f"{username}@chak.org"

        # Keep existing password if one exists; else generate a fresh unique one
        password = None
        if not FORCE:
            password = existing_excel_passwords.get(email) or existing_json_passwords.get(
                email
            )
        if not password:
            password = gen_password()
            while password in used_passwords:
                password = gen_password()
            used_passwords.add(password)

        users[email] = {
            "email": email,
            "password": password,
            "role": "staff",
            "name": name,
            "facility": facility,
            "county": county,
            "jobTitle": designation,
            "phone": phone,
            "idNumber": id_num,
        }

        row[username_col - 1].value = email
        row[password_col - 1].value = password

    try:
        wb.save(XLSX)
        print(f"Excel updated: {len(users)} staff credentials saved to {XLSX}")
    except PermissionError:
        print(f"ERROR: {XLSX} is locked — please CLOSE it in Excel and re-run.")

    # ── Supervisors (not in Excel — created from the org structure) ──
    # Facility → County map derived from the staff roster (each facility sits in one county)
    facility_county: dict[str, str] = {}
    for u in users.values():
        if u["role"] == "staff" and u["facility"] and u["county"]:
            facility_county.setdefault(u["facility"], u["county"])

    def password_for(email: str) -> str:
        """Reuse existing password unless --force; otherwise mint a unique one."""
        if not FORCE:
            p = existing_json_passwords.get(email)
            if p:
                return p
        p = gen_password()
        while p in used_passwords:
            p = gen_password()
        used_passwords.add(p)
        return p

    # Facility In-Charges: one per unique facility, linked to its county (and thus its County Rep)
    facilities = sorted({u["facility"] for u in users.values()})
    for facility in facilities:
        base = slugify(f"incharge {facility}")
        email = f"{base}@chak.org"
        users[email] = {
            "email": email,
            "password": password_for(email),
            "role": "facility_incharge",
            "name": "Facility In-Charge",
            "facility": facility,
            "county": facility_county.get(facility, ""),
            "jobTitle": "Facility In-Charge",
            "phone": "",
        }

    # County Reps: one per unique county
    counties = sorted({u["county"] for u in users.values() if u["county"]})
    for county in counties:
        base = slugify(f"countyrep {county}")
        email = f"{base}@chak.org"
        users[email] = {
            "email": email,
            "password": password_for(email),
            "role": "county_rep",
            "name": "County Representative",
            "facility": "",
            "county": county,
            "jobTitle": "County Representative",
            "phone": "",
        }

    # Program HR
    users["hr@chak.org"] = {
        "email": "hr@chak.org",
        "password": password_for("hr@chak.org"),
        "role": "program_hr",
        "name": "Program HR",
        "facility": "",
        "county": "",
        "jobTitle": "Human Resource Officer",
        "phone": "",
    }

    # Demo accounts (used in earlier testing)
    demo = {
        "staff@chak.org": {"role": "staff", "name": "Staff", "facility": "Main Facility", "county": "Nairobi", "jobTitle": "Medical Officer"},
        "incharge@chak.org": {"role": "facility_incharge", "name": "Facility In-Charge", "facility": "Main Facility", "county": "Nairobi", "jobTitle": "Facility In-Charge"},
        "countyrep@chak.org": {"role": "county_rep", "name": "County Rep", "facility": "", "county": "Embu", "jobTitle": "County Representative"},
    }
    for email, info in demo.items():
        if email not in users:
            users[email] = {
                "email": email,
                "password": password_for(email),
                "phone": "",
                "idNumber": "",
                **info,
            }

    # Final uniqueness sweep (shouldn't normally fire, but be safe)
    seen: dict[str, str] = {}
    for email, u in users.items():
        p = u["password"]
        if p in seen:
            new = gen_password()
            while new in seen or new in used_passwords:
                new = gen_password()
            seen[email] = new
            users[email]["password"] = new
        else:
            seen[email] = p

    # Role counts
    roles = Counter(u["role"] for u in users.values())
    print("Role breakdown:")
    for role, count in roles.items():
        print(f"  {role}: {count}")

    os.makedirs("data", exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(users, f, indent=2, ensure_ascii=False)
    print(f"Users written to {OUT_JSON}: {len(users)} total")

    passwords = [u["password"] for u in users.values()]
    print(f"Unique passwords: {len(set(passwords))} / {len(passwords)}")

    # Print sample credentials
    print("\nSample logins:")
    sample = [u for u in users.values() if u["role"] == "staff"][:3]
    for u in sample:
        print(f"  {u['email']} / {u['password']} ({u['name']} - {u['facility']})")
    for u in users.values():
        if u["role"] == "facility_incharge":
            print(f"  {u['email']} / {u['password']} (In-Charge: {u['facility']})")
            break
    for u in users.values():
        if u["role"] == "county_rep":
            print(f"  {u['email']} / {u['password']} (County Rep: {u['county']})")
            break
    print(f"  hr@chak.org / {users['hr@chak.org']['password']} (Program HR)")


if __name__ == "__main__":
    main()
